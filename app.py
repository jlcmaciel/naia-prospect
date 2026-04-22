"""
Naia Prospect — Plataforma M&A
Arquivo único autocontido — sem dependência de módulos cacheados
"""
import streamlit as st
import pandas as pd
import requests
import sqlite3, os, tempfile, networkx as nx, io
from datetime import datetime
from functools import lru_cache

# ── Paleta Naia ───────────────────────────────────────────────────────────────
VERDE_ESCURO = "#1B5E20"
VERDE_MEDIO  = "#2E7D32"
VERDE_CLARO  = "#7BC67E"
BRANCO       = "#FFFFFF"
FUNDO        = "#0F1F0F"
FUNDO_CARD   = "#162616"
BORDA        = "#2E7D32"

# ── Exportação Excel formatada ────────────────────────────────────────────────
def to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Dados", title: str = "") -> bytes:
    """Gera um .xlsx formatado com cabeçalho Naia, larguras automáticas e formatação de números."""
    from openpyxl import Workbook
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Estilos
    HDR_FILL  = PatternFill("solid", fgColor="1B5E20")   # verde escuro Naia
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
    TITLE_FONT= Font(bold=True, color="1B5E20", size=13)
    ALT_FILL  = PatternFill("solid", fgColor="EAF4EA")   # verde bem claro (linhas pares)
    THIN_SIDE = Side(style="thin", color="C8E6C9")
    BORDER    = Border(left=THIN_SIDE, right=THIN_SIDE,
                       top=THIN_SIDE, bottom=THIN_SIDE)
    CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=False)
    LEFT      = Alignment(horizontal="left",   vertical="center", wrap_text=False)

    # Linha de título (linha 1)
    label = title or sheet_name
    ws.cell(row=1, column=1, value=f"Naia Prospect — {label}")
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = LEFT
    ws.row_dimensions[1].height = 20

    # Cabeçalho das colunas (linha 2)
    cols = list(df.columns)
    for ci, col in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=ci, value=str(col))
        cell.fill   = HDR_FILL
        cell.font   = HDR_FONT
        cell.border = BORDER
        cell.alignment = CENTER
    ws.row_dimensions[2].height = 18

    # Detecta colunas numéricas para formatação
    NUM_FMT_MONEY = '#,##0.0'   # valores em R$ MM
    NUM_FMT_INT   = '#,##0'
    NUM_FMT_DATE  = 'DD/MM/YYYY'

    money_cols = {ci+1 for ci, c in enumerate(cols)
                  if any(kw in c.lower() for kw in ("r$ mm","capital","fat","ebitda","receita","lucro","usd","export"))}
    date_cols  = {ci+1 for ci, c in enumerate(cols)
                  if any(kw in c.lower() for kw in ("fundacao","data","date","ano"))}

    # Dados
    for ri, row in enumerate(df.itertuples(index=False), start=3):
        fill = ALT_FILL if (ri % 2 == 0) else None
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = BORDER
            if fill:
                cell.fill = fill
            if ci in money_cols:
                cell.number_format = NUM_FMT_MONEY
                cell.alignment = CENTER
            elif ci in date_cols:
                cell.number_format = NUM_FMT_DATE
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT

    # Auto-ajuste de largura das colunas
    for ci, col in enumerate(cols, start=1):
        col_letter = get_column_letter(ci)
        max_len = max(
            len(str(col)),
            *[len(str(row[ci-1])) for row in df.itertuples(index=False)]
        ) if len(df) > 0 else len(str(col))
        # Limita entre 10 e 45 caracteres
        ws.column_dimensions[col_letter].width = max(10, min(max_len + 2, 45))

    # Congela linha do cabeçalho
    ws.freeze_panes = "A3"

    # Auto-filter no cabeçalho
    last_col = get_column_letter(len(cols))
    ws.auto_filter.ref = f"A2:{last_col}2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ── Login ─────────────────────────────────────────────────────────────────────
def _check_login():
    """Tela de login simples. Usuários e senhas ficam no secrets.toml."""
    # Se não há usuários configurados, pula o login (modo dev local)
    usuarios = st.secrets.get("usuarios", {})
    if not usuarios:
        return True

    if st.session_state.get("_autenticado"):
        return True

    # CSS específico para a tela de login (fundo escuro, card centralizado)
    st.markdown(f"""
    <style>
      .stApp {{background-color:{FUNDO}}}
      .login-card {{
        max-width:400px; margin:6rem auto 0; padding:2.5rem 2rem;
        background:{FUNDO_CARD}; border:1px solid {BORDA};
        border-radius:14px; box-shadow:0 4px 32px #00000060;
      }}
      .login-logo {{
        text-align:center; font-size:2.8rem; margin-bottom:.3rem;
      }}
      .login-title {{
        text-align:center; color:{BRANCO}; font-family:Georgia,serif;
        font-size:1.5rem; margin:0 0 .2rem;
      }}
      .login-sub {{
        text-align:center; color:{VERDE_CLARO}; font-size:.8rem;
        margin:0 0 2rem;
      }}
    </style>
    <div class="login-card">
      <div class="login-logo">🌿</div>
      <p class="login-title">Naia Prospect</p>
      <p class="login-sub">Inteligência M&A · Uso interno</p>
    </div>
    """, unsafe_allow_html=True)

    with st.form("login_form"):
        usuario = st.text_input("Usuário", placeholder="seu.nome")
        senha   = st.text_input("Senha", type="password", placeholder="••••••••")
        entrar  = st.form_submit_button("Entrar →", type="primary", use_container_width=True)

    if entrar:
        senha_correta = usuarios.get(usuario, "")
        if senha_correta and senha == senha_correta:
            st.session_state["_autenticado"] = True
            st.session_state["_usuario"]     = usuario
            st.rerun()
        else:
            st.error("Usuário ou senha incorretos.")
    return False

st.set_page_config(page_title="Naia Prospect", layout="wide", page_icon="🌿")
st.markdown(f"""<style>
  .stApp{{background-color:{FUNDO};color:{BRANCO}}}
  section[data-testid="stSidebar"]{{background-color:{FUNDO_CARD};border-right:1px solid {BORDA}}}
  h1,h2,h3,h4{{color:{BRANCO};font-family:Georgia,serif}}
  p,label,div[data-testid="stMarkdownContainer"] p{{color:#D4E8D4}}
  div.stButton>button[kind="primary"]{{background-color:{VERDE_MEDIO}!important;color:{BRANCO}!important;border:none!important;border-radius:6px!important;font-weight:600!important;padding:.5rem 2rem!important}}
  div.stButton>button[kind="primary"]:hover{{background-color:{VERDE_CLARO}!important;color:{FUNDO}!important}}
  div.stButton>button{{border:1px solid {VERDE_CLARO}!important;color:{VERDE_CLARO}!important;background:transparent!important;border-radius:6px!important}}
  div[data-baseweb="input"]>div,div[data-baseweb="select"]>div{{background-color:{FUNDO_CARD}!important;border-color:{BORDA}!important;color:{BRANCO}!important;border-radius:6px!important}}
  div[data-testid="metric-container"]{{background-color:{FUNDO_CARD};border:1px solid {BORDA};border-radius:8px;padding:1rem}}
  div[data-testid="metric-container"] label{{color:{VERDE_CLARO}!important;font-size:.8rem}}
  div[data-testid="metric-container"] div[data-testid="stMetricValue"]{{color:{BRANCO}!important;font-size:1.6rem}}
  button[data-baseweb="tab"]{{color:#A5C8A5!important}}
  button[data-baseweb="tab"][aria-selected="true"]{{color:{VERDE_CLARO}!important;border-bottom:2px solid {VERDE_CLARO}!important}}
  div[data-testid="stDataFrame"]{{border:1px solid {BORDA};border-radius:8px}}
  div[data-testid="stRadio"] label{{color:#A5C8A5!important;font-size:.95rem}}
  details summary{{color:{VERDE_CLARO}!important}}
  details{{border:1px solid {BORDA}!important;border-radius:8px;background:{FUNDO_CARD}}}
  ::-webkit-scrollbar{{width:6px}}
  ::-webkit-scrollbar-thumb{{background:{VERDE_MEDIO};border-radius:3px}}
</style>""", unsafe_allow_html=True)

# ── Verificação de login ──────────────────────────────────────────────────────
if not _check_login():
    st.stop()

# ── Header ────────────────────────────────────────────────────────────────────
c1, c2 = st.columns([1,10])
c1.markdown(f'<div style="background:{VERDE_MEDIO};border-radius:50%;width:52px;height:52px;display:flex;align-items:center;justify-content:center;margin-top:4px"><span style="font-size:1.6rem">🌿</span></div>', unsafe_allow_html=True)
c2.markdown(f'<h1 style="margin:0;color:{BRANCO}">Naia Prospect</h1><p style="margin:0;color:{VERDE_CLARO};font-size:.85rem">Inteligência M&A · CVM · RFB 60M CNPJs · Comex Stat</p>', unsafe_allow_html=True)
st.markdown("<hr style='margin:.5rem 0 1rem 0'>", unsafe_allow_html=True)

# ── Sidebar ───────────────────────────────────────────────────────────────────
_user = st.session_state.get("_usuario", "")
st.sidebar.markdown(f'<div style="text-align:center;padding:1rem 0 .5rem"><span style="color:{VERDE_CLARO};font-size:1.1rem;font-weight:700;letter-spacing:2px">NAIA CAPITAL</span><br><span style="color:#A5C8A5;font-size:.75rem">M&A Intelligence Platform</span></div>', unsafe_allow_html=True)
if _user:
    st.sidebar.markdown(f'<div style="text-align:center;padding:.2rem 0 .4rem"><span style="color:#7BC67E;font-size:.8rem">👤 {_user}</span></div>', unsafe_allow_html=True)
    if st.sidebar.button("Sair", key="_logout"):
        st.session_state.pop("_autenticado", None)
        st.session_state.pop("_usuario", None)
        st.rerun()
st.sidebar.markdown("---")
modo = st.sidebar.radio("Selecione o modo", ["🏦 Dado Real (CVM)", "🔍 Busca Ampla (60M CNPJs)", "🌎 Exportadores (Comex)", "📋 Pipeline Naia"])

# ── BigQuery (inline, sem cache de módulo) ────────────────────────────────────
def _bq_creds():
    if "gcp_service_account" in st.secrets:
        from google.oauth2 import service_account
        return service_account.Credentials.from_service_account_info(dict(st.secrets["gcp_service_account"]))
    return None

def bq_query(sql: str) -> pd.DataFrame:
    import pandas_gbq
    project = st.secrets.get("BILLING_PROJECT", "")
    if not project:
        return pd.DataFrame()
    return pandas_gbq.read_gbq(sql, project_id=project, credentials=_bq_creds())

def buscar_rfb(uf=None, cnae="", porte="", fat_min_mm=0, fat_max_mm=0,
               ano_min=1900, ano_max=2026, limite=500,
               nome="", so_com_email=False) -> pd.DataFrame:
    """
    Busca empresas na RFB via BigQuery.
    Filtros de faturamento são aplicados em Python pós-query (capital_social
    não é proxy confiável de faturamento no Brasil — muitas empresas lucrativas
    têm capital baixo). O BigQuery faz apenas os filtros estruturais.
    """
    where = ["est.situacao_cadastral = '2'"]

    # Filtros estruturais (aplicados direto no BigQuery)
    if uf:
        if isinstance(uf, list) and len(uf) == 1:
            where.append(f"est.sigla_uf = '{uf[0]}'")
        elif isinstance(uf, list) and len(uf) > 1:
            ufs_str = ",".join([f"'{u}'" for u in uf])
            where.append(f"est.sigla_uf IN ({ufs_str})")
        elif isinstance(uf, str) and uf:
            where.append(f"est.sigla_uf = '{uf}'")
    if cnae:
        where.append(f"STARTS_WITH(CAST(est.cnae_fiscal_principal AS STRING), '{cnae}')")
    if porte:
        where.append(f"emp.porte = '{porte}'")
    if ano_min > 1900:
        where.append(f"EXTRACT(YEAR FROM est.data_inicio_atividade) >= {ano_min}")
    if ano_max < 2026:
        where.append(f"EXTRACT(YEAR FROM est.data_inicio_atividade) <= {ano_max}")
    if nome:
        where.append(f"UPPER(emp.razao_social) LIKE '%{nome.upper().strip()}%'")
    if so_com_email:
        where.append("est.email IS NOT NULL AND est.email != ''")

    # Quando há filtro de fat, amplia o limite para garantir candidatos suficientes
    bq_limite = limite * 6 if (fat_min_mm > 0 or fat_max_mm > 0) else limite

    sql = f"""
    SELECT est.cnpj, emp.razao_social, est.nome_fantasia,
           est.cnae_fiscal_principal AS cnae, est.sigla_uf AS uf,
           est.id_municipio AS municipio, est.cep, est.email,
           CONCAT(IFNULL(est.ddd_1,''), IFNULL(est.telefone_1,'')) AS telefone,
           emp.capital_social, emp.porte,
           est.data_inicio_atividade AS fundacao, est.cnpj_basico
    FROM `basedosdados.br_me_cnpj.estabelecimentos` est
    JOIN `basedosdados.br_me_cnpj.empresas` emp
      ON est.cnpj_basico = emp.cnpj_basico
    WHERE est.ano = 2024 AND est.mes = 4
      AND emp.ano = 2024 AND emp.mes = 4
      AND {' AND '.join(where)}
    ORDER BY emp.capital_social DESC
    LIMIT {bq_limite}"""
    return bq_query(sql)

def buscar_socios(cnpj_basicos: tuple) -> pd.DataFrame:
    if not cnpj_basicos: return pd.DataFrame()
    ids = ",".join([f"'{c}'" for c in cnpj_basicos])
    sql = f"""
    SELECT s.cnpj_basico, s.nome_socio, s.qualificacao_socio, e.razao_social
    FROM `basedosdados.br_me_cnpj.socios` s
    JOIN `basedosdados.br_me_cnpj.empresas` e ON s.cnpj_basico = e.cnpj_basico
    WHERE s.ano = 2024 AND s.mes = 4 AND e.ano = 2024 AND e.mes = 4
      AND s.cnpj_basico IN ({ids})"""
    return bq_query(sql)

# ── Estimador de faturamento ──────────────────────────────────────────────────
# Multiplicador receita/capital por setor (benchmarks mercado brasileiro)
# fat_estimado = capital_social × multiplicador
MULT_FAT = {
    "10":8,"11":6,"13":4,"17":5,"19":3,"20":4,"21":5,"22":4,"24":4,
    "26":3,"28":3,"29":4,"41":4,"46":6,"47":6,"49":4,"61":3,
    "62":3,"64":2,"68":2,"86":4,"85":3
}
MARGEM_EBITDA = {
    "10":.12,"11":.25,"13":.10,"17":.22,"19":.18,"20":.20,"21":.28,
    "22":.15,"24":.18,"26":.20,"28":.14,"29":.12,"41":.10,"46":.07,
    "47":.08,"49":.15,"61":.30,"62":.20,"64":.45,"68":.35,"86":.22,"85":.18,
}
def fat_estimado(capital, cnae):
    p = str(cnae or "")[:2]
    mult = MULT_FAT.get(p, 3)
    return max(0, float(capital or 0)) * mult
def ebitda_estimado(fat, cnae):
    p = str(cnae or "")[:2]
    return fat * MARGEM_EBITDA.get(p, 0.15)
def porte_label(fat):
    if fat < 4_800_000:    return "Micro (<R$4,8M)"
    if fat < 30_000_000:   return "Pequena (R$4,8M–30M)"
    if fat < 300_000_000:  return "Média (R$30M–300M)"
    if fat < 1_000_000_000:return "Média-grande (R$300M–1B)"
    return "Grande (>R$1B)"

def score_ma(row, fat_max_ref, fat_min_ref):
    """Score M&A 0–100 baseado em faturamento, maturidade, EBITDA e contato disponível."""
    score = 0
    fat   = float(row.get("fat_est", 0) or 0)
    ebitda= float(row.get("ebitda_est", 0) or 0)
    # 1. Faturamento (0–35 pts) — normalizado dentro dos resultados
    fat_range = max(fat_max_ref - fat_min_ref, 1)
    score += 35 * min((fat - fat_min_ref) / fat_range, 1.0)
    # 2. Margem EBITDA estimada (0–25 pts)
    margem = ebitda / fat if fat > 0 else 0
    score += min(margem / 0.40, 1.0) * 25
    # 3. Maturidade (anos desde fundação) — empresas com 5–30 anos pontuam mais (0–25 pts)
    try:
        fund = pd.to_datetime(row.get("fundacao", None), errors="coerce")
        anos = (datetime.now() - fund).days / 365 if pd.notna(fund) else 0
        if   anos >= 5:  score += 25
        elif anos >= 2:  score += 15
        elif anos >= 1:  score += 5
    except: pass
    # 4. Contato disponível (0–15 pts)
    if str(row.get("email","") or "").strip(): score += 8
    if str(row.get("telefone","") or "").strip(): score += 7
    return round(min(score, 100))

# ── Buscas Salvas SQLite ──────────────────────────────────────────────────────
import json as _json
def _db_searches():
    c = sqlite3.connect(DB)
    c.execute("""CREATE TABLE IF NOT EXISTS buscas_salvas(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT UNIQUE,
        filtros TEXT,
        criado_em TEXT,
        ultima_exec TEXT)""")
    return c
def busca_salvar(nome, filtros: dict):
    c = _db_searches()
    c.execute("INSERT INTO buscas_salvas(nome,filtros,criado_em,ultima_exec) VALUES(?,?,?,?) "
              "ON CONFLICT(nome) DO UPDATE SET filtros=excluded.filtros,ultima_exec=excluded.ultima_exec",
              (nome, _json.dumps(filtros, ensure_ascii=False), datetime.now().isoformat(), datetime.now().isoformat()))
    c.commit(); c.close()
def busca_listar():
    c = _db_searches()
    try: rows = pd.read_sql("SELECT * FROM buscas_salvas ORDER BY ultima_exec DESC", c)
    except: rows = pd.DataFrame()
    c.close(); return rows
def busca_deletar(nome):
    c = _db_searches(); c.execute("DELETE FROM buscas_salvas WHERE nome=?", (nome,)); c.commit(); c.close()

# ── Enriquecimento ────────────────────────────────────────────────────────────
@lru_cache(maxsize=500)
def enriquecer(cnpj):
    c = "".join(filter(str.isdigit, str(cnpj)))
    try:
        r = requests.get(f"https://brasilapi.com.br/api/cnpj/v1/{c}", timeout=12)
        if r.ok: return r.json()
    except: pass
    try:
        r = requests.get(f"https://minhareceita.org/{c}", timeout=12)
        if r.ok: return r.json()
    except: pass
    return None

@lru_cache(maxsize=500)
def geocode(cep):
    c = "".join(filter(str.isdigit, str(cep or "")))
    if len(c) != 8: return None
    try:
        r = requests.get(f"https://cep.awesomeapi.com.br/json/{c}", timeout=6)
        if r.ok:
            d = r.json()
            if d.get("lat"): return float(d["lat"]), float(d["lng"])
    except: pass
    return None

# ── Pipeline SQLite ───────────────────────────────────────────────────────────
DB = os.path.join(os.path.dirname(__file__), "cache", "pipeline.db")
os.makedirs(os.path.dirname(DB), exist_ok=True)
STATUS_OPTS = ["Target identificado","Pesquisa em andamento","Contato inicial feito",
               "Reunião agendada","NDA assinado","Diligência","Proposta enviada",
               "Negociação","Descartado","Fechado"]
def _db():
    c = sqlite3.connect(DB)
    c.execute("CREATE TABLE IF NOT EXISTS pipeline(cnpj TEXT PRIMARY KEY,razao_social TEXT,status TEXT,responsavel TEXT,nota TEXT,atualizado_em TEXT)")
    return c
def pipe_upsert(cnpj, razao, status, resp, nota):
    c=_db(); c.execute("INSERT INTO pipeline VALUES(?,?,?,?,?,?) ON CONFLICT(cnpj) DO UPDATE SET razao_social=excluded.razao_social,status=excluded.status,responsavel=excluded.responsavel,nota=excluded.nota,atualizado_em=excluded.atualizado_em",(cnpj,razao,status,resp,nota,datetime.now().isoformat())); c.commit(); c.close()
def pipe_listar():
    c=_db(); df=pd.read_sql("SELECT * FROM pipeline ORDER BY atualizado_em DESC",c); c.close(); return df
def pipe_obter(cnpj):
    c=_db(); row=c.execute("SELECT * FROM pipeline WHERE cnpj=?",(cnpj,)).fetchone(); c.close()
    return dict(zip(["cnpj","razao_social","status","responsavel","nota","atualizado_em"],row)) if row else {}
def pipe_remover(cnpj):
    c=_db(); c.execute("DELETE FROM pipeline WHERE cnpj=?",(cnpj,)); c.commit(); c.close()

# ── CVM ───────────────────────────────────────────────────────────────────────
CACHE_DIR = os.path.join(os.path.dirname(__file__), "cache")
def cvm_buscar(query="", receita_min=0, ano=2023):
    import zipfile, io
    cache = os.path.join(CACHE_DIR, f"cvm_dre_{ano}.parquet")
    if os.path.exists(cache):
        df = pd.read_parquet(cache)
    else:
        url = f"https://dados.cvm.gov.br/dados/CIA_ABERTA/DOC/DFP/DADOS/dfp_cia_aberta_{ano}.zip"
        r = requests.get(url, timeout=120)
        if not r.ok: return pd.DataFrame()
        with zipfile.ZipFile(io.BytesIO(r.content)) as z:
            target = f"dfp_cia_aberta_DRE_con_{ano}.csv"
            if target not in z.namelist(): return pd.DataFrame()
            with z.open(target) as f:
                df = pd.read_csv(f, sep=";", encoding="latin1", decimal=",")
        df = df[df["ORDEM_EXERC"].str.contains("ÚLTIMO", na=False)]
        keys = {"3.03":"receita_liquida","3.05":"resultado_bruto","3.11":"lucro_liquido"}
        base = df[df["CD_CONTA"].isin(list(keys.keys()))].copy()
        base["metrica"] = base["CD_CONTA"].map(keys)
        piv = base.pivot_table(index=["CNPJ_CIA","DENOM_CIA"], columns="metrica", values="VL_CONTA", aggfunc="first").reset_index()
        piv.columns.name = None
        for col in ["receita_liquida","resultado_bruto","lucro_liquido"]:
            if col in piv.columns: piv[col] = pd.to_numeric(piv[col], errors="coerce")
        piv["ebitda_proxy"] = piv.get("resultado_bruto", 0)
        piv["ano"] = ano
        df = piv.rename(columns={"CNPJ_CIA":"cnpj","DENOM_CIA":"razao_social"})
        df.to_parquet(cache, index=False)
    if query:
        df = df[df["razao_social"].str.upper().str.contains(query.upper(), na=False)]
    if receita_min > 0:
        df = df[pd.to_numeric(df["receita_liquida"], errors="coerce").fillna(0) >= receita_min]
    return df.sort_values("receita_liquida", ascending=False, na_position="last")

# ═══════════════════════════════════════════════════════════════════════════════
# MODO 1 — CVM
# ═══════════════════════════════════════════════════════════════════════════════
if modo.startswith("🏦"):
    st.markdown(f'<h2>Companhias Abertas <span style="background:{VERDE_CLARO};color:{FUNDO};padding:2px 10px;border-radius:12px;font-size:.75rem;font-weight:600">DADO REAL</span></h2><p style="color:#A5C8A5;font-size:.85rem">Receita e resultado reais das DFPs entregues à CVM (~470 empresas)</p>', unsafe_allow_html=True)
    c1,c2,c3 = st.columns([2,3,2])
    ano = c1.selectbox("Ano base", [2023,2022,2021,2020])
    busca = c2.text_input("🔎 Buscar por nome", placeholder="ex: Ambev, Petrobras...")
    rec_min = c3.number_input("Receita mín. (R$ MM)", 0, value=0, step=50)
    if st.button("Buscar na CVM →", type="primary"):
        with st.spinner("Carregando dados da CVM..."):
            df = cvm_buscar(busca, rec_min*1e6, ano)
            st.session_state["cvm_df"] = df
    if "cvm_df" in st.session_state:
        df = st.session_state["cvm_df"]
        if df.empty:
            st.warning("Nenhuma empresa encontrada.")
        else:
            s = df.copy()
            s["Receita (R$ MM)"] = (pd.to_numeric(s.get("receita_liquida",0),errors="coerce")/1e6).round(1)
            s["EBITDA (R$ MM)"]  = (pd.to_numeric(s.get("ebitda_proxy",0),errors="coerce")/1e6).round(1)
            s["Lucro (R$ MM)"]   = (pd.to_numeric(s.get("lucro_liquido",0),errors="coerce")/1e6).round(1)
            m1,m2,m3 = st.columns(3)
            m1.metric("Empresas", len(s))
            m2.metric("Maior receita", f"R$ {pd.to_numeric(s['receita_liquida'],errors='coerce').max()/1e9:.1f}B")
            m3.metric("Receita total", f"R$ {pd.to_numeric(s['receita_liquida'],errors='coerce').sum()/1e9:.0f}B")
            t1,t2 = st.tabs(["📊 Resultados","🏢 Ficha"])
            with t1:
                cols = [c for c in ["cnpj","razao_social","Receita (R$ MM)","EBITDA (R$ MM)","Lucro (R$ MM)","ano"] if c in s.columns]
                st.dataframe(s[cols], use_container_width=True, height=460)
                st.download_button("⬇️ Exportar Excel", to_excel_bytes(s[cols], "CVM", f"CVM {ano}"),
                                   f"naia_cvm_{ano}.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            with t2:
                cnpjs = s["cnpj"].dropna().astype(str).tolist()
                alvo = st.selectbox("Empresa", cnpjs)
                if alvo:
                    info = enriquecer(alvo)
                    if info:
                        st.markdown(f'<div style="background:{FUNDO_CARD};border:1px solid {BORDA};border-radius:10px;padding:1.2rem"><h3 style="color:{VERDE_CLARO};margin:0 0 .5rem">{info.get("razao_social","")}</h3><p style="color:#ccc;margin:0;font-size:.85rem">{info.get("cnae_fiscal_descricao","")}</p><p style="color:#aaa;margin:.3rem 0 0;font-size:.8rem">📍 {info.get("municipio","")}/{info.get("uf","")} &nbsp;|&nbsp; 📞 {info.get("ddd_telefone_1","") or "-"} &nbsp;|&nbsp; ✉️ {info.get("email","") or "-"}</p></div>', unsafe_allow_html=True)
                    with st.expander("➕ Adicionar ao Pipeline"):
                        s2=st.selectbox("Status",STATUS_OPTS); r=st.text_input("Responsável"); n=st.text_area("Notas")
                        if st.button("Salvar", type="primary"):
                            pipe_upsert(alvo, info.get("razao_social","") if info else "", s2, r, n)
                            st.success("Salvo!")

# ═══════════════════════════════════════════════════════════════════════════════
# MODO 2 — RFB 60M
# ═══════════════════════════════════════════════════════════════════════════════
elif modo.startswith("🔍"):
    st.markdown(f'<h2>Busca Ampla <span style="background:#F59E0B;color:{FUNDO};padding:2px 10px;border-radius:12px;font-size:.75rem;font-weight:600">ESTIMADO ±40%</span></h2><p style="color:#A5C8A5;font-size:.85rem">60 milhões de CNPJs da Receita Federal com faturamento estimado por setor</p>', unsafe_allow_html=True)

    gcp_ok = bool(st.secrets.get("BILLING_PROJECT",""))
    if not gcp_ok:
        st.error("⚠️ Configure BILLING_PROJECT no secrets.toml")
        st.stop()

    SETORES = {"":"Todos os setores","10":"🍎 Alimentos","11":"🍺 Bebidas","13":"👕 Têxtil",
               "17":"📄 Papel/Celulose","19":"⛽ Petróleo","20":"🧪 Química","21":"💊 Farmacêutica",
               "22":"🔧 Borracha/Plástico","24":"⚙️ Metalurgia","26":"💻 Eletrônicos",
               "28":"🏭 Máquinas","29":"🚗 Veículos","41":"🏗️ Construção","46":"📦 Atacado",
               "47":"🛒 Varejo","49":"🚛 Transporte","61":"📡 Telecom","62":"💻 TI/Software",
               "64":"🏦 Financeiro","68":"🏢 Imobiliário","86":"🏥 Saúde","85":"🎓 Educação"}
    TODAS_UFS = ["SP","RJ","MG","RS","PR","SC","BA","GO","DF","PE","CE","ES","PA","MT","MS","AM","MA","RN","PB","AL","SE","PI","RO","AC","AP","RR","TO"]

    # ── Buscas Salvas ─────────────────────────────────────────────────────────
    buscas_df = busca_listar()
    if not buscas_df.empty:
        with st.expander(f"📂 Buscas salvas ({len(buscas_df)})", expanded=False):
            for _, bs in buscas_df.iterrows():
                bc1, bc2, bc3 = st.columns([4, 2, 1])
                bc1.markdown(f"**{bs['nome']}**")
                bc2.caption(f"Última exec: {str(bs.get('ultima_exec',''))[:10]}")
                col_run, col_del = bc3.columns(2)
                if col_run.button("▶", key=f"run_{bs['nome']}", help="Executar busca"):
                    filtros = _json.loads(bs["filtros"])
                    st.session_state["_busca_carregada"] = filtros
                    st.rerun()
                if col_del.button("🗑", key=f"del_{bs['nome']}", help="Excluir"):
                    busca_deletar(bs["nome"]); st.rerun()

    # pré-carrega filtros se uma busca salva foi acionada
    _pre = st.session_state.pop("_busca_carregada", {})

    # ── Filtros ───────────────────────────────────────────────────────────────
    st.markdown(f"<p style='color:{VERDE_CLARO};font-weight:600;margin-bottom:.3rem'>Setor & Empresa</p>", unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    setor       = c1.selectbox("Setor", list(SETORES.keys()), format_func=lambda k: SETORES[k],
                                index=list(SETORES.keys()).index(_pre.get("setor","")) if _pre.get("setor","") in SETORES else 0)
    cnae_manual = c1.text_input("Ou CNAE manual", value=_pre.get("cnae_manual",""), placeholder="ex: 62, 10, 86")
    cnae        = cnae_manual.strip() or setor
    nome_busca  = c2.text_input("🔎 Buscar por nome da empresa", value=_pre.get("nome_busca",""), placeholder="ex: Ambev, Grupo XP...")
    porte       = c2.selectbox("Porte", ["","01","03","05"],
                                format_func=lambda k: {"":"Todos","01":"Micro","03":"EPP","05":"Médio/Grande"}[k],
                                index=["","01","03","05"].index(_pre.get("porte","")) if _pre.get("porte","") in ["","01","03","05"] else 0)

    st.markdown(f"<p style='color:{VERDE_CLARO};font-weight:600;margin:.8rem 0 .3rem'>Localização</p>", unsafe_allow_html=True)
    ufs_sel = st.multiselect("Estados (pode selecionar vários)", TODAS_UFS, default=_pre.get("ufs_sel",[]))

    st.markdown(f"<p style='color:{VERDE_CLARO};font-weight:600;margin:.8rem 0 .3rem'>Tamanho & Período</p>", unsafe_allow_html=True)
    c4,c5,c6,c7,c8 = st.columns(5)
    fat_min    = c4.number_input("Fat. mín (R$ MM)", 0, value=int(_pre.get("fat_min",0)), step=10)
    fat_max    = c5.number_input("Fat. máx (R$ MM)", 0, value=int(_pre.get("fat_max",0)), step=50)
    ano_ini    = c6.number_input("Fundada após", 1900, 2026, int(_pre.get("ano_ini",1900)))
    ano_fim    = c7.number_input("Fundada até", 1900, 2026, int(_pre.get("ano_fim",2026)))
    limite     = c8.slider("Máx resultados", 50, 2000, int(_pre.get("limite",500)), 50)

    st.markdown(f"<p style='color:{VERDE_CLARO};font-weight:600;margin:.8rem 0 .3rem'>Filtros adicionais</p>", unsafe_allow_html=True)
    fa1, fa2, fa3 = st.columns(3)
    so_email       = fa1.checkbox("📧 Apenas com e-mail cadastrado", value=_pre.get("so_email", False))
    excluir_pipe   = fa2.checkbox("🚫 Excluir empresas já no Pipeline", value=_pre.get("excluir_pipe", False))
    ordenar_score  = fa3.checkbox("⭐ Ordenar por Score M&A", value=_pre.get("ordenar_score", True), help="Ordena os resultados pelo score de atratividade M&A calculado automaticamente")

    ba1, ba2 = st.columns([2, 1])
    buscar_btn = ba1.button("🔍  Buscar empresas →", type="primary")

    with ba2.container():
        nome_salvar = st.text_input("💾 Nome para salvar busca", placeholder="ex: TI SP acima 50MM", label_visibility="collapsed")
        salvar_btn  = st.button("💾 Salvar esta busca", use_container_width=True)

    if salvar_btn:
        if nome_salvar:
            busca_salvar(nome_salvar, {
                "setor":setor,"cnae_manual":cnae_manual,"nome_busca":nome_busca,
                "porte":porte,"ufs_sel":ufs_sel,"fat_min":fat_min,"fat_max":fat_max,
                "ano_ini":ano_ini,"ano_fim":ano_fim,"limite":limite,
                "so_email":so_email,"excluir_pipe":excluir_pipe,"ordenar_score":ordenar_score
            })
            st.success(f"✅ Busca '{nome_salvar}' salva!"); st.rerun()
        else:
            st.warning("Digite um nome antes de salvar.")

    if buscar_btn:
        with st.spinner("Consultando base da Receita Federal..."):
            try:
                df_raw = buscar_rfb(
                    uf=ufs_sel or None, cnae=cnae, porte=porte,
                    fat_min_mm=fat_min, fat_max_mm=fat_max,
                    ano_min=ano_ini, ano_max=ano_fim,
                    limite=limite, nome=nome_busca, so_com_email=so_email
                )
                n_bq = len(df_raw)
                if df_raw.empty:
                    st.warning("Nenhuma empresa encontrada no BigQuery. Tente ampliar os filtros.")
                else:
                    df = df_raw.copy()
                    df["fat_est"]    = df.apply(lambda r: fat_estimado(r.get("capital_social",0), r.get("cnae","")), axis=1)
                    df["ebitda_est"] = df.apply(lambda r: ebitda_estimado(r.get("fat_est",0), r.get("cnae","")), axis=1)
                    df["porte_est"]  = df["fat_est"].apply(porte_label)
                    n_antes = len(df)
                    if fat_min > 0: df = df[df["fat_est"] >= fat_min * 1_000_000]
                    if fat_max > 0: df = df[df["fat_est"] <= fat_max * 1_000_000]
                    df = df.head(limite)   # respeita o limite após filtro
                    n_depois = len(df)
                    # Score M&A
                    if not df.empty:
                        fat_vals = df["fat_est"]
                        fat_min_r, fat_max_r = fat_vals.min(), fat_vals.max()
                        df["Score M&A"] = df.apply(lambda r: score_ma(r, fat_max_r, fat_min_r), axis=1)
                    # Excluir do pipeline
                    if excluir_pipe:
                        pipe_cnpjs = set(pipe_listar()["cnpj"].astype(str).tolist())
                        df = df[~df["cnpj"].astype(str).isin(pipe_cnpjs)]
                    st.session_state["rfb_df"] = df
                    st.session_state["rfb_ordenar_score"] = ordenar_score
                    if df.empty:
                        st.warning(f"⚠️ BigQuery retornou {n_bq} empresas, mas o filtro de faturamento removeu todas.")
                        # diagnóstico: mostra faixa real de fat_est encontrada
                        df_diag = df_raw.copy()
                        df_diag["fat_est"] = df_diag.apply(lambda r: fat_estimado(r.get("capital_social",0), r.get("cnae","")), axis=1)
                        fat_min_real = df_diag["fat_est"].min() / 1_000_000
                        fat_max_real = df_diag["fat_est"].max() / 1_000_000
                        st.info(f"💡 Faturamento estimado das empresas encontradas: **R$ {fat_min_real:.0f}MM** a **R$ {fat_max_real:.0f}MM**. "
                                f"Seu filtro era R$ {fat_min}MM a R$ {fat_max}MM. Ajuste a faixa.")
                    else:
                        st.success(f"✅ {n_depois} empresas encontradas (BigQuery: {n_bq} → após filtro fat.: {n_depois})")
            except Exception as e:
                st.error(f"Erro: {e}"); st.code(str(e))

    if "rfb_df" in st.session_state:
        df = st.session_state["rfb_df"]
        _ord_score = st.session_state.get("rfb_ordenar_score", True)
        if _ord_score and "Score M&A" in df.columns:
            df = df.sort_values("Score M&A", ascending=False)

        m1,m2,m3,m4,m5 = st.columns(5)
        m1.metric("Empresas", len(df))
        m2.metric("Fat. médio estim.", f"R$ {df['fat_est'].mean()/1e6:.1f}MM" if len(df) else "-")
        m3.metric("EBITDA médio estim.", f"R$ {df['ebitda_est'].mean()/1e6:.1f}MM" if len(df) else "-")
        m4.metric("Capital médio", f"R$ {df['capital_social'].mean()/1e6:.1f}MM" if len(df) else "-")
        m5.metric("Score médio M&A", f"{df['Score M&A'].mean():.0f}/100" if "Score M&A" in df.columns and len(df) else "-")

        t1,t2,t3,t4 = st.tabs(["📊 Resultados","🗺️ Mapa","🕸️ Grupos Econômicos","🏢 Ficha"])
        with t1:
            s = df.copy()
            s["Fat. Est. (R$MM)"]   = (s["fat_est"]/1e6).round(1)
            s["EBITDA Est. (R$MM)"] = (s["ebitda_est"]/1e6).round(1)
            s["Capital (R$MM)"]     = (s["capital_social"]/1e6).round(1)
            cols = ["Score M&A","razao_social","cnae","uf","municipio","Fat. Est. (R$MM)","EBITDA Est. (R$MM)","Capital (R$MM)","porte_est","email","fundacao"]
            show = s[[c for c in cols if c in s.columns]]
            st.dataframe(show, use_container_width=True, height=460)
            export_df = show.copy()
            st.download_button("⬇️ Exportar Excel", to_excel_bytes(export_df, "Busca Ampla", "Busca Ampla — RFB"),
                               "naia_prospect.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        with t2:
            st.caption("Geocodificando top 50 resultados...")
            am = df.head(50).copy()
            am["coords"] = am["cep"].apply(geocode)
            am["lat"] = am["coords"].apply(lambda x: x[0] if x else None)
            am["lon"] = am["coords"].apply(lambda x: x[1] if x else None)
            mdf = am.dropna(subset=["lat","lon"])
            if len(mdf): st.map(mdf[["lat","lon"]], zoom=4)
            else: st.info("Nenhum CEP geocodificável nesta amostra.")

        with t3:
            st.caption("Detecta sócios em comum — revela grupos econômicos ocultos.")
            topn = st.slider("Top N empresas", 10, 200, 50)
            if st.button("Gerar grafo", type="primary"):
                try:
                    from pyvis.network import Network
                    import streamlit.components.v1 as components
                    basicos = tuple(df["cnpj_basico"].head(topn).dropna().unique())
                    socios = buscar_socios(basicos)
                    G = nx.Graph()
                    for _, row in socios.iterrows():
                        G.add_node(f"E:{str(row['razao_social'])[:35]}", color=VERDE_MEDIO, size=20)
                        G.add_node(f"S:{str(row['nome_socio'])[:35]}", color=VERDE_CLARO, size=12)
                        G.add_edge(f"E:{str(row['razao_social'])[:35]}", f"S:{str(row['nome_socio'])[:35]}")
                    grupos = [n for n,d in G.degree() if n.startswith("S:") and d>1]
                    st.metric("Grupos econômicos detectados", len(grupos))
                    net = Network(height="600px", width="100%", bgcolor=FUNDO, font_color=BRANCO)
                    net.from_nx(G)
                    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".html")
                    net.save_graph(tmp.name)
                    with open(tmp.name, encoding="utf-8") as f: components.html(f.read(), height=620)
                    os.unlink(tmp.name)
                    if grupos:
                        st.subheader("Sócios em múltiplas empresas")
                        for g in grupos[:20]:
                            ems = list(G.neighbors(g))
                            st.markdown(f"• **{g[2:]}** → {len(ems)} empresas: {', '.join(e[2:] for e in ems[:5])}")
                except Exception as e:
                    st.error(str(e))

        with t4:
            if len(df):
                opts = df["razao_social"].fillna("(sem nome)").astype(str).tolist()
                idx  = st.selectbox("Empresa", range(len(opts)), format_func=lambda i: opts[i])
                linha = df.iloc[idx]
                alvo  = str(linha["cnpj"])
                c1,c2,c3 = st.columns(3)
                c1.metric("Fat. estimado", f"R$ {linha['fat_est']/1e6:.1f}MM")
                c2.metric("EBITDA estimado", f"R$ {linha['ebitda_est']/1e6:.1f}MM")
                c3.metric("Porte estimado", str(linha.get("porte_est","-")))
                info = enriquecer(alvo)
                if info:
                    st.markdown(f'<div style="background:{FUNDO_CARD};border:1px solid {BORDA};border-radius:10px;padding:1.2rem"><h3 style="color:{VERDE_CLARO};margin:0 0 .5rem">{info.get("razao_social","")}</h3><p style="color:#ccc;margin:0;font-size:.85rem">{info.get("cnae_fiscal_descricao","")}</p><p style="color:#aaa;margin:.3rem 0 0;font-size:.8rem">📍 {info.get("municipio","")}/{info.get("uf","")} &nbsp;|&nbsp; 📞 {info.get("ddd_telefone_1","") or "-"} &nbsp;|&nbsp; ✉️ {info.get("email","") or "-"}</p><p style="color:{VERDE_CLARO};margin:.5rem 0 0;font-size:.85rem">💰 Fat. est: <b>R$ {linha["fat_est"]/1e6:.1f}MM</b> &nbsp;|&nbsp; EBITDA est: <b>R$ {linha["ebitda_est"]/1e6:.1f}MM</b></p></div>', unsafe_allow_html=True)
                    if info.get("qsa"):
                        with st.expander("👥 Sócios (QSA)"):
                            st.dataframe(pd.DataFrame(info["qsa"]), use_container_width=True)
                with st.expander("➕ Adicionar ao Pipeline"):
                    s2=st.selectbox("Status",STATUS_OPTS,key="rfb_s"); r=st.text_input("Responsável",key="rfb_r"); n=st.text_area("Notas",key="rfb_n")
                    if st.button("Salvar no pipeline →", type="primary", key="rfb_b"):
                        pipe_upsert(alvo, str(linha["razao_social"]), s2, r, n)
                        st.success("✅ Adicionado ao pipeline!")

# ═══════════════════════════════════════════════════════════════════════════════
# MODO 3 — Comex
# ═══════════════════════════════════════════════════════════════════════════════
elif modo.startswith("🌎"):
    st.markdown(f'<h2>Exportações por Estado <span style="background:{VERDE_CLARO};color:{FUNDO};padding:2px 10px;border-radius:12px;font-size:.75rem;font-weight:600">DADO REAL</span></h2><p style="color:#A5C8A5;font-size:.85rem">Volume exportado em US$ por UF e capítulo NCM — Comex Stat / MDIC</p>', unsafe_allow_html=True)

    DETALHE_OPTS = {
        "state":   "🗺️ Por Estado (UF)",
        "chapter": "📦 Por Capítulo NCM",
        "section": "🏭 Por Seção",
    }
    c1,c2,c3,c4 = st.columns(4)
    ano_i   = c1.number_input("Ano início", 2010, 2025, 2023)
    ano_f   = c2.number_input("Ano fim",    2010, 2025, 2024)
    detalhe = c3.selectbox("Agrupar por", list(DETALHE_OPTS.keys()), format_func=lambda k: DETALHE_OPTS[k])
    fluxo   = c4.selectbox("Fluxo", ["export","import"], format_func=lambda k: {"export":"Exportação","import":"Importação"}[k])

    st.markdown(f"<p style='color:{VERDE_CLARO};font-weight:600;margin:.6rem 0 .2rem'>Filtros opcionais</p>", unsafe_allow_html=True)
    f1,f2 = st.columns(2)
    uf_c  = f1.selectbox("Estado (filtro)", ["","SP","RJ","MG","RS","PR","SC","BA","GO","DF","ES","PE","CE","AM","PA","MT","GO"])
    ncm_c = f2.text_input("Capítulo NCM", placeholder="ex: 02 = Carnes, 84 = Máquinas, 27 = Petróleo")

    if st.button("Buscar dados Comex →", type="primary"):
        with st.spinner("Consultando Comex Stat / MDIC..."):
            payload = {
                "flow": fluxo,
                "monthDetail": False,
                "period": {"from": f"{ano_i}-01", "to": f"{ano_f}-12"},
                "filters": [],
                "details": [detalhe],
                "metrics": ["metricFOB", "metricKG"]
            }
            if uf_c:  payload["filters"].append({"filter":"state",   "values":[uf_c]})
            if ncm_c: payload["filters"].append({"filter":"chapter",  "values":[ncm_c.zfill(2)]})
            try:
                r = requests.post("https://api-comexstat.mdic.gov.br/general", json=payload, timeout=30)
                if r.ok:
                    data = r.json().get("data",{}).get("list",[])
                    if data:
                        df = pd.DataFrame(data)
                        # normaliza coluna de valor
                        fob_col = next((c for c in df.columns if "FOB" in c.upper() or "fob" in c), None)
                        kg_col  = next((c for c in df.columns if "KG"  in c.upper() or "kg"  in c), None)
                        df["Exportação (US$MM)"] = pd.to_numeric(df[fob_col], errors="coerce").div(1e6).round(2) if fob_col else 0
                        df["Peso (mil ton)"]     = pd.to_numeric(df[kg_col],  errors="coerce").div(1e6).round(1) if kg_col  else 0
                        df = df.sort_values("Exportação (US$MM)", ascending=False)
                        # renomeia coluna de agrupamento
                        label_col = {"state":"Estado","chapter":"Capítulo NCM","section":"Seção"}.get(detalhe, detalhe)
                        if detalhe in df.columns:
                            df = df.rename(columns={detalhe: label_col})
                        m1,m2,m3 = st.columns(3)
                        m1.metric("Registros", len(df))
                        m2.metric("Total exportado", f"US$ {df['Exportação (US$MM)'].sum():.0f}MM")
                        m3.metric("Maior", f"{df[label_col].iloc[0] if label_col in df.columns else '-'}")
                        show_cols = [c for c in [label_col,"year","Exportação (US$MM)","Peso (mil ton)"] if c in df.columns]
                        st.dataframe(df[show_cols], use_container_width=True, height=460)
                        st.download_button("⬇️ Exportar Excel",
                                           to_excel_bytes(df[show_cols], "Comex", "Exportações Comex Stat"),
                                           "naia_comex.xlsx",
                                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    else:
                        st.warning("Sem dados para os filtros selecionados. Tente ampliar o período ou remover filtros.")
                else:
                    st.warning(f"API Comex retornou erro {r.status_code}. Tente novamente em alguns segundos.")
                    st.code(r.text[:300])
            except Exception as e:
                st.error(f"Erro ao consultar Comex Stat: {e}")

    # ── Cruzamento Comex × RFB ────────────────────────────────────────────────
    st.markdown("---")
    st.markdown(f"<h3 style='color:{BRANCO}'>🔗 Prospectar empresas do setor exportador</h3><p style='color:#A5C8A5;font-size:.85rem'>Identifique o setor pelo Comex acima, depois busque os CNPJs correspondentes na aba Busca Ampla.</p>", unsafe_allow_html=True)

    CRUZAMENTO = [
        ("🌾 Agronegócio / Carnes",      "02, 04, 10", "10", "Alimentos — CNAE 10xx"),
        ("🌲 Papel & Celulose",           "47, 48",     "17", "Papel/Celulose — CNAE 17xx"),
        ("⛽ Petróleo & Derivados",       "27",         "19", "Petróleo — CNAE 19xx"),
        ("🧪 Química & Fertilizantes",    "28, 31",     "20", "Química — CNAE 20xx"),
        ("💊 Farmacêutica",               "30",         "21", "Farmacêutica — CNAE 21xx"),
        ("⚙️ Metalurgia & Minério",       "26, 72",     "24", "Metalurgia — CNAE 24xx"),
        ("💻 Eletrônicos & Semicond.",    "85",         "26", "Eletrônicos — CNAE 26xx"),
        ("🏭 Máquinas & Equipamentos",    "84",         "28", "Máquinas — CNAE 28xx"),
        ("🚗 Veículos & Autopeças",       "87, 84",     "29", "Veículos — CNAE 29xx"),
        ("✈️ Aeronáutica",               "88",         "30", "Aeronáutica — CNAE 30xx"),
        ("🚛 Transporte & Log.",          "89, 86",     "49", "Transporte — CNAE 49xx"),
    ]

    cols_header = st.columns([3, 2, 3, 2])
    cols_header[0].markdown(f"<b style='color:{VERDE_CLARO}'>Setor exportador</b>", unsafe_allow_html=True)
    cols_header[1].markdown(f"<b style='color:{VERDE_CLARO}'>NCM (Comex)</b>", unsafe_allow_html=True)
    cols_header[2].markdown(f"<b style='color:{VERDE_CLARO}'>Buscar empresas por</b>", unsafe_allow_html=True)
    cols_header[3].markdown(f"<b style='color:{VERDE_CLARO}'>CNAE</b>", unsafe_allow_html=True)

    for setor, ncm_ref, cnae_cod, cnae_label in CRUZAMENTO:
        c1, c2, c3, c4 = st.columns([3, 2, 3, 2])
        c1.markdown(f"<span style='color:{BRANCO}'>{setor}</span>", unsafe_allow_html=True)
        c2.markdown(f"<code style='background:{FUNDO_CARD};color:{VERDE_CLARO};padding:2px 6px;border-radius:4px'>{ncm_ref}</code>", unsafe_allow_html=True)
        c3.markdown(f"<span style='color:#A5C8A5;font-size:.9rem'>{cnae_label}</span>", unsafe_allow_html=True)
        if c4.button(f"Buscar →", key=f"cross_{cnae_cod}"):
            st.session_state["_cnae_sugerido"] = cnae_cod
            st.session_state["_setor_sugerido"] = setor
            st.info(f"💡 Vá para **🔍 Busca Ampla** e use o CNAE `{cnae_cod}` para encontrar essas empresas.")

    if "_cnae_sugerido" in st.session_state:
        cnae_s = st.session_state["_cnae_sugerido"]
        setor_s = st.session_state.get("_setor_sugerido","")
        st.markdown(f"""
        <div style='background:{FUNDO_CARD};border:1px solid {BORDA};border-radius:10px;padding:1rem;margin-top:.5rem'>
          <p style='color:{VERDE_CLARO};margin:0;font-weight:600'>💡 Próximo passo sugerido</p>
          <p style='color:{BRANCO};margin:.3rem 0 0'>Acesse <b>🔍 Busca Ampla</b> no menu lateral e filtre por CNAE <code>{cnae_s}</code> para prospectar empresas do setor <b>{setor_s}</b>.</p>
        </div>
        """, unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# MODO 4 — Pipeline
# ═══════════════════════════════════════════════════════════════════════════════
else:
    st.markdown(f'<h2>Pipeline Naia</h2><p style="color:#A5C8A5;font-size:.85rem">Prospecções ativas da equipe de M&A</p>', unsafe_allow_html=True)
    df = pipe_listar()
    if df.empty:
        st.info("📭 Nenhuma prospecção salva. Use os outros modos para adicionar alvos.")
    else:
        m1,m2,m3,m4 = st.columns(4)
        m1.metric("Total", len(df))
        m2.metric("Em negociação", len(df[df["status"].isin(["Proposta enviada","Negociação","Diligência"])]))
        m3.metric("Em andamento",  len(df[df["status"].isin(["Pesquisa em andamento","Contato inicial feito","Reunião agendada","NDA assinado"])]))
        m4.metric("Fechados/Descartados", len(df[df["status"].isin(["Fechado","Descartado"])]))
        st.dataframe(df, use_container_width=True, height=400)
        st.download_button("⬇️ Exportar Excel", to_excel_bytes(df, "Pipeline", "Pipeline Naia M&A"),
                           "pipeline_naia.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        st.subheader("Editar prospecção")
        alvo = st.selectbox("Empresa", df["cnpj"].tolist(), format_func=lambda c: f"{df[df['cnpj']==c]['razao_social'].values[0]} ({c})")
        if alvo:
            p = pipe_obter(alvo)
            idx = STATUS_OPTS.index(p.get("status","Target identificado")) if p.get("status") in STATUS_OPTS else 0
            ns=st.selectbox("Status",STATUS_OPTS,index=idx); nr=st.text_input("Responsável",p.get("responsavel","")); nn=st.text_area("Notas",p.get("nota",""))
            b1,b2 = st.columns(2)
            if b1.button("Atualizar →", type="primary"):
                pipe_upsert(alvo, p.get("razao_social",""), ns, nr, nn); st.success("Atualizado!"); st.rerun()
            if b2.button("Remover"):
                pipe_remover(alvo); st.rerun()

# ── Footer ────────────────────────────────────────────────────────────────────
st.sidebar.markdown("---")
gcp_status = "🟢 BigQuery conectado" if st.secrets.get("BILLING_PROJECT","") else "🔴 Sem BigQuery"
st.sidebar.markdown(f'<div style="text-align:center"><span style="color:#7BC67E;font-size:.8rem">{gcp_status}</span><br><span style="color:#4a7a4a;font-size:.7rem">© Naia Capital — uso interno</span></div>', unsafe_allow_html=True)
